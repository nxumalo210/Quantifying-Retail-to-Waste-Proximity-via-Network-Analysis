# this code is going to be extracting the city name and address of all Macy's store locations across the state of NY
#
# import openpyxl
# from openpyxl import load_workbook
# import time
# import pandas as pd
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.chrome.service import Service as ChromeService
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.common.exceptions import TimeoutException
#
# def store_collector():
#     chrome_options = Options()
#     chrome_options.add_argument("--headless")
#     chrome_options.add_argument("--no-sandbox")  # Helps in some environments
#     chrome_options.add_argument("--disable-dev-shm-usage")
#
#     driver = webdriver.Chrome()
#     driver.get("https://www.macys.com/stores/ny/albany")
#
#     try:
#         stores_list_container = WebDriverWait(driver, 10).until(EC.presence_of_element_located(By.ID, "map-list-wrap"))
#         stores = driver.find_elements(By.CLASS_NAME, "map-list")
#     except TimeoutException:
#         print("Timed out waiting for list of stores")
#
#     data = []
#
#     for store in stores:
#         try:
#             store_name = stores[0].find_elements(By.CLASS_NAME, "location-name is-block").text
#
#             address = stores.find_elements(By.CLASS_NAME, "address").text
#
#             data.append({"Store Name": store_name, "Address": address})
#         except Exception as e:
#             print("Error while extracting data from store: ", e)
#             continue
#
#     driver.quit()
#
#     df = pd.DataFrame(data)
#     file_path = "/Users/neonxumalo/Documents/Data Science Practice Files/GIS Practice/Independent Projects/Macys_Store_Location_Data_NY.csv"
#
#     try:
#         book = load_workbook(file_path)
#         sheet_name = "Macys_Store_Location_Data_NY" if "Macys_Store_Location_Data_NY" in book.sheetnames else \
#         book.sheetnames[1]
#         sheet = book[sheet_name]
#         start_row = sheet.max_row
#
#         with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#             df.to_csv(file_path, mode='a', index=False, header=not os.path.exists(file_path))
#             print(f"\n✅ Appended {len(df)} rows to existing CSV file:\n{file_path}")
#
#     except Exception as e:
#         print("Error while extracting data from file: ", e)
#
#
# if __name__ == "__main__":
#     store_collector()

# Use undetected_chromedriver library
# ... (imports) ...

import os
import pandas as pd
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook  # Needed for the ExcelWriter engine


def store_collector():
    """
    Navigates to a Macy's store page, scrapes store name and address,
    and returns a list of dictionaries containing the data.
    """
    # --- Driver Setup ---
    options = uc.ChromeOptions()
    options.add_argument("--disable-gpu")
    options.page_load_strategy = 'eager'
    driver = uc.Chrome(options=options, use_subprocess=True, headless=False)
    print("Undetected ChromeDriver initiated. Page load strategy set to 'eager'.")

    data = []

    try:
        # --- Navigation ---
        driver.get("https://www.macys.com/stores/ny/yorktownheights/")
        print("Page requested. Waiting for initial DOM load.")

        # --- Handle Cookie Consent Banner ---
        try:
            print("Checking for cookie banner (ID: onetrust-accept-btn-handler)...")
            accept_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
            )
            accept_button.click()
            print("Cookie banner accepted.")
        except TimeoutException:
            print("No cookie banner found or timed out waiting for it.")

        # --- Handle the Signup Banner if it blocks interactions ---
        # NOTE: You MUST replace 'close-button-class-name' with the actual class name
        try:
            print("Checking for sign-up pop-up to close it...")
            close_signup = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CLASS_NAME, "close-button-class-name"))
            )
            close_signup.click()
            print("Sign-up banner closed.")
        except TimeoutException:
            print("No sign-up banner found or could not close it.")

        # --- Wait for Main Content using the Corrected ID ---
        store_list_container = WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.ID, "section-browse-locations"))
        )
        print("Container 'section-browse-locations' found and is visible.")

        # --- Scrape Data using Corrected Selectors ---
        # Assuming these selectors are correct for the target URL
        store_items = store_list_container.find_elements(By.CLASS_NAME, "map-list-item-wrap")
        print(f"Found {len(store_items)} store items using 'map-list-item-wrap' class.")

        for item in store_items:
            store_name_element = item.find_element(By.CLASS_NAME, "location-name")
            address_element = item.find_element(By.CLASS_NAME, "address")
            store_name = store_name_element.text.strip()
            address = address_element.text.strip().replace('\n', ', ')
            print(f"Scraped: Name='{store_name}', Address='{address}'")
            data.append({"Store Name": store_name, "Address": address})

    except TimeoutException:
        print("Timed out waiting for main container. The page might still be loading.")
        driver.save_screenshot("eager_load_timeout_screenshot.png")

    finally:
        if driver:
            driver.quit()
            print("Browser closed.")

    # Return the scraped data list
    return data


def save_to_excel(data_list):
    """
    Processes the list of data into a DataFrame and saves/appends to an Excel file.
    """
    df = pd.DataFrame(data_list)
    if df.empty:
        print("\n❌ DataFrame is empty. No data to save.")
        return

    file_path = "/Users/neonxumalo/Documents/Data Science Practice Files/GIS Practice/Independent Projects/Macys_Store_Location_Data_NY.xlsx"
    sheet_name = "Macys_Store_Location_Data_NY"

    # Ensure the directory exists
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    file_exists = os.path.exists(file_path)

    try:
        if not file_exists:
            # Create New File with Headers
            df.to_excel(file_path, sheet_name=sheet_name, index=False, header=True)
            print(f"\n✅ Created new workbook and wrote {len(df)} rows to:\n{file_path}")

        else:
            # Append to existing file using openpyxl engine and 'overlay' mode
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Determine the next empty row
                startrow = writer.book[sheet_name].max_row
                # Append the data without writing the header again
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
            print(f"\n✅ Appended {len(df)} new rows to existing workbook file:\n{file_path}")

    except Exception as e:
        print(f"Error while writing data to file: {e}")


# This block is now correctly outside and runs the main logic
if __name__ == "__main__":
    # 1. Scrape the data and get the result back as a list
    scraped_data = store_collector()

    # 2. Process and save the data to Excel
    save_to_excel(scraped_data)
